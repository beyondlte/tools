import os
import csv
import openpyxl
from pprint import pprint
from lxml import etree as ET 

def convertCsv2Xlsx(csvFile):
    csvDir = os.path.dirname(csvFile)
    csvName = os.path.basename(csvFile)
    fileName = csvName.split('.')[0]

    wb = openpyxl.Workbook()
    ws = wb.active

    f = open(csvFile)
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)
    f.close()

    xlsxFilename = os.path.join(csvDir, fileName+'.xlsx')
    wb.save(xlsxFilename)
    return xlsxFilename

def checkSeq(xlsxFile, rework):
    wb = openpyxl.load_workbook(xlsxFile)
    ws = wb.get_sheet_by_name('Sheet')
    maxRow = ws.max_row
    maxCol = ws.max_column

    rown = 1

    while rown <= maxRow:
        reworkName = ws.cell(row=rown, column=2).value
        if rework == reworkName:
            seqList = []
            for cl in range(4, maxCol):
                cellVal = ws.cell(row=rown, column = cl).value
                if cellVal != 'skip':
                    opName = ws.cell(row=2, column = cl).value
                    opNameMod = ws.cell(row=3, column = cl).value
                    if opNameMod != None:
                        # seqList.append({"{}.{}".format(opName, opNameMod):str(cellVal)})
                        seqList.append("{}.{}".format(opName, opNameMod))
                    else:
                        #seqList.append({"{}".format(opName):str(cellVal)})
                        seqList.append("{}".format(opName))
        rown += 1

    print("="*50)
    print(rework)
    """
    print("-"*50)
    pprint(seqList)
    print("="*50)
    """
    return seqList

def getSeqByOperationName(root):
    try:
        seqDict = [] 
        for elem in  root.iter('Test'):
            # allChildren is a list
            # each element in this allChildren list is lxml.Etree._Element 
            # which has attrib, base, tag, tail, text, values etc.
            allChildren = elem.getchildren()
            specTag = 0
            currentOpname = ''
            if allChildren == []:
                continue
            else:
                for chm in allChildren:
                    if chm.tag is ET.Comment:
                        continue
                    elif 'arg' in chm.tag and 'OPERATION_NAME' in chm.get('name'):
                        if elem.get('visible_mode') == None:
                            # seqDict.append({chm.text:elem.get('visible_mode')})
                            seqDict.append("{}".format(chm.text[1:-1]))
                          
        return seqDict 
    except Exception as e:
        traceback.print_exc()

def getTests(root):
    try:
        seqDict = [] 
        for elem in root.iter('Test'):
            allChildren = elem.getchildren()
            if allChildren == []:
                continue
            else:
                if elem.get('visible_mode') == None:
                    seqDict.append(elem.get('name'))
        return seqDict
    except Execption as e:
        traceback.print_exc()



def getSeqFromTestXml(testxml):
    tree = ET.parse(testxml)
    root = tree.getroot()
    testSeqDict = getSeqByOperationName(root)
    # pprint(testSeqDict)
    return testSeqDict


def getTestsFromTestXml(testxml):
    tree = ET.parse(testxml)
    root = tree.getroot()
    filebase = testxml.split('.')[0] 
    print "------------- Test Suite of {} ---------------".format(os.path.basename(testxml))
    with open(filebase + '.txt', 'w') as f:
        for t in getTests(root):
            print t
            f.write(t + '\n')

def compareTestSeq(xlsxFile, rework, xmlFile):

    csvSeqList = checkSeq(xlsxFile, rework)
    xmlSeqList = getSeqFromTestXml(xmlFile)

    initIndex = csvSeqList.index('init')

    if csvSeqList[initIndex:] == xmlSeqList:
        print 'checking test seq: Pass'
    else:
        print 'checking test seq: Fail'

masterCsv = r'C:\SystemNPI\Projects\SP1098\DteInterface\SP1098.2_P2_Limited_Launch\AIM\AIM_sequence_of_operations_style_3221.csv'
masterXml = convertCsv2Xlsx(masterCsv)

psmTestXml = r'C:\SystemNPI\Projects\SP1098\PSM2\trunk\sp1098_psm_tests_unittest.xml'
motorTestXml = r'C:\SystemNPI\Projects\SP1098\PSM2\trunk\sp1098_motorpack_tests.xml'
IOTestXml = r'C:\SystemNPI\Projects\SP1098\PSM2\trunk\sp1098_insertionmotor_tests.xml'

# compareTestSeq(masterXml, 'first_time_cal_after_both_modules',  psmTestXml)
compareTestSeq(masterXml, 'motorpack_module', motorTestXml) 
"""
compareTestSeq(masterXml, 'insertion_motor_module', IOTestXml) 

getTestsFromTestXml(psmTestXml)
getTestsFromTestXml(IOTestXml)
"""
getTestsFromTestXml(motorTestXml)

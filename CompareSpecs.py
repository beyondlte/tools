from lxml import etree as ET 
from pprint import pprint
import traceback
import exceptions
import re

from openpyxl import load_workbook


count = 0
def getAllSpecs(elem, lstOut = None):
    global count

    if lstOut == None:
        lstOut = []

    for chm in elem.getchildren():
        if chm.tag is not ET.Comment and 'arg' in chm.tag:
            count += 1
            lstOut.append(chm.get('name'))
        else:
            getAllSpecs(chm, lstOut)
    return lstOut 


def getOperationName(root):
    try:
        specDict = {} 
        for elem in  root.iter('Test'):
            # allChildren is a list
            # each element in this allChildren list is lxml.Etree._Element 
            # which has attrib, base, tag, tail, text, values etc.
            allChildren = elem.getchildren()
            specTag = 0
            currentOpname = ''
            if allChildren == []:
                # print 0
                continue
            else:
                for chm in allChildren:
                    if chm.tag is ET.Comment:
                        continue
                    elif 'arg' in chm.tag and 'OPERATION_NAME' in chm.get('name'):
                        specDict[chm.text] = {} 
                        currentOpname = chm.text
                    elif 'spec' in chm.tag:
                        if currentOpname == '':
                            continue
                        else:
                            # specDict[currentOpname] should not be a list, should also be a dict
                            # {name: {tag:dof1, lsl:1...}}
                            dKey = '{}_{}'.format(chm.get('name'), chm.get('tag'))
                            specDict[currentOpname][dKey] = { \
                                    'name':chm.get('name'), 
                                    'tag':chm.get('tag', None), 
                                    'type':chm.get('type'), 
                                    'lsl':chm.get('lsl'), 
                                    'usl':chm.get('usl'),
                                    'flag_sap':chm.get('flag_sap'),
                                    'units':chm.get('units', None), 
                                    'reference':chm.get('reference', None)}
                          
        return specDict 
    except Exception as e:
        traceback.print_exc()


def compareDicts(xmlFile1, dict1, xmlFile2, dict2):
    print "-"*160
    print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
            'xml file name',
            'test name',
            'spec name',
            'tag',
            'units',
            'lsl',
            'usl',
            'flag_sap',
            'reference')
    print "-"*160

    for em in dict1:
        if dict1[em] == {}:
            print("{} has no specs".format(em))
            
        elif dict1[em] != {}:
            spDict1 = dict1[em]
            if em not in dict2:
                print("!!! {} not in {}".format(em, xmlFile2))
                continue 
            spDict2 = dict2[em]
            for sp1 in spDict1:
                if sp1 in spDict2:
                    if spDict1[sp1]['tag'] == spDict2[sp1]['tag']:
                        if spDict1[sp1]['type'] == 'abovelsl':
                            if float(spDict1[sp1]['lsl']) != float(spDict2[sp1]['lsl']):
                                print '{:>40}, {:>10}:, {:>10}, {:>10}, {:>10}'.format(\
                                        xmlFile1,
                                        em,
                                        sp1,
                                        spDict1[sp1]['tag'],
                                        spDict1[sp1]['lsl'])
                                        
                                print '{:>40}, {:>24}:, {:>10}, {:>10}, {:>10} '.format(\
                                        xmlFile2,
                                        em,
                                        sp1,
                                        spDict2[sp1]['tag'],
                                        spDict2[sp1]['lsl'])
                                print "-"*80

                        elif spDict1[sp1]['flag_sap'] == '1' and spDict2[sp1]['flag_sap'] != '1': #(spDict2[sp1]['reference'] != 'spDict1[sp1]098_PSM' or spDict2[sp1]['flag_sap'] != 1):
                            print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile1,
                                    em,
                                    spDict1[sp1]['name'],
                                    spDict1[sp1]['tag'],
                                    spDict1[sp1]['units'],
                                    spDict1[sp1]['lsl'],
                                    spDict1[sp1]['usl'],
                                    spDict1[sp1]['flag_sap'])
                            print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile2,
                                    em,
                                    spDict2[sp1]['name'],
                                    spDict2[sp1]['tag'],
                                    spDict2[sp1]['units'],
                                    spDict2[sp1]['lsl'],
                                    spDict2[sp1]['usl'],
                                    spDict2[sp1]['flag_sap'],
                                    spDict2[sp1].get('reference', None))

                        elif spDict1[sp1]['lsl'] == None:
                            continue
                        elif float(spDict1[sp1]['lsl']) != float(spDict2[sp1]['lsl']) or float(spDict1[sp1]['usl']) != float(spDict2[sp1]['usl']) or \
                                spDict1[sp1]['units'] != spDict2[sp1]['units']:

                            print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile1,
                                    em,
                                    spDict1[sp1]['name'],
                                    spDict1[sp1]['tag'],
                                    spDict1[sp1]['units'],
                                    spDict1[sp1]['lsl'],
                                    spDict1[sp1]['usl'])
                            print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile2,
                                    em,
                                    spDict2[sp1]['name'],
                                    spDict2[sp1]['tag'],
                                    spDict2[sp1]['units'],
                                    spDict2[sp1]['lsl'],
                                    spDict2[sp1]['usl'])
                            print "-"*130

                else:
                    print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile1,
                                    em,
                                    spDict1[sp1]['name'],
                                    spDict1[sp1]['tag'],
                                    spDict1[sp1]['units'],
                                    spDict1[sp1]['lsl'],
                                    spDict1[sp1]['usl'])
                    print '{:>40}, {:>25}:, {:>25} '.format(\
                                    xmlFile2,
                                    em,
                                    '---NA---'
                                    )
                    print "-"*130 



# compare specs in master.xml and tests.xml
def compareDicts_old(xmlFile1, dict1, xmlFile2, dict2):
    print "="*130
    print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
            'xml file name',
            'test name',
            'spec name',
            'tag',
            'lsl',
            'usl',
            'flag_sap',
            'units',
            'reference')
    print "="*130

    for em in dict1:
        print "em = ", em
        if dict1[em] != []:
            spList1 = dict1[em]
            spList2 = dict2[em]
            pprint(spList1)
            pprint(spList2)
            break
            for sp1 in spList1:
                # for each sp1, check if there's a matched sp2
                for sp2 in spList2:
                    if sp1['tag'] == sp2['tag'] and sp1['name'] == sp2['name']:
                        if sp1['type'] == 'abovelsl': 
                            if float(sp1['lsl']) != float(sp2['lsl']):
                                print '{:>40}, {:>10}:, {:>10}, {:>10}, {:>10}'.format(\
                                        xmlFile1,
                                        em,
                                        sp1['name'],
                                        sp1['tag'],
                                        sp1['lsl'])
                                        
                                print '{:>40}, {:>24}:, {:>10}, {:>10}, {:>10} '.format(\
                                        xmlFile2,
                                        em,
                                        sp2['name'],
                                        sp2['tag'],
                                        sp2['lsl'])
                                print "-"*80
                            continue
                        elif sp1['flag_sap'] == '1' and sp2['flag_sap'] != '1': #(sp2['reference'] != 'SP1098_PSM' or sp2['flag_sap'] != 1):
                            print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile1,
                                    em,
                                    sp1['name'],
                                    sp1['tag'],
                                    sp1['units'],
                                    sp1['lsl'],
                                    sp1['usl'],
                                    sp1['flag_sap'])
                            print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile2,
                                    em,
                                    sp2['name'],
                                    sp2['tag'],
                                    sp2['units'],
                                    sp2['lsl'],
                                    sp2['usl'],
                                    sp2['flag_sap'],
                                    sp2.get('reference', None))

                        elif sp1['lsl'] == None:
                            continue
                        elif float(sp1['lsl']) != float(sp2['lsl']) or float(sp1['usl']) != float(sp2['usl']) or \
                                sp1['units'] != sp2['units']:

                            print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile1,
                                    em,
                                    sp1['name'],
                                    sp1['tag'],
                                    sp1['units'],
                                    sp1['lsl'],
                                    sp1['usl'])
                            print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile2,
                                    em,
                                    sp2['name'],
                                    sp2['tag'],
                                    sp2['units'],
                                    sp2['lsl'],
                                    sp2['usl'])
                            print "-"*130
                    elif sp1['name'] != sp2['name']:
                        print '{:>40}, {:>25}:, {:>25}, {:>10}, {:>10}, {:>10}, {:>10}'.format(\
                                    xmlFile1,
                                    em,
                                    sp1['name'],
                                    sp1['tag'],
                                    sp1['units'],
                                    sp1['lsl'],
                                    sp1['usl'])


def compareTwoXmls(xmlFile1, xmlFile2):
    # create spec dicts by "operation name"
    tree1 = ET.parse(xmlFile1)
    root1 = tree1.getroot()
    specDict1 = getOperationName(root1)

    tree2 = ET.parse(xmlFile2)
    root2 = tree2.getroot()
    specDict2 = getOperationName(root2)

    # pprint(specDict1["'friction.motor_module'"])
    print("+"*80)
    print('Comparing {} and {}'.format(xmlFile1, xmlFile2))
    compareDicts(xmlFile1, specDict1, xmlFile2, specDict2)
    print '\n'
    print('Comparing {} and {}'.format(xmlFile2, xmlFile1))
    compareDicts(xmlFile2, specDict2, xmlFile1, specDict1)
    print("+"*80)


# compare specs in xml file and xlsx file
def compareXmlXlsx(xmlFile, xlsxFile):
    wb2 = load_workbook(xlsxFile)
    ws2 = wb2.get_sheet_by_name('ResultsSheet')
    maxRow2 = ws2.max_row
    row2 = 5

    # parse xml file and save specs in structure 
    #       sap_uid: name, tag, units, type, lsl, usl
    tree1 = ET.parse(xmlFile)
    root1 = tree1.getroot()
    specDictXml = {} 
    for specElem in  root1.iter('spec'):
        specUid = specElem.get('sap_uid')
        if specUid == None:
            continue
        else:
            specDictXml[specElem.get('sap_uid')] =  [specElem.get('name'), specElem.get('tag'), 
                                     specElem.get('type'), specElem.get('lsl'),
                                     specElem.get('usl'), specElem.get('units'),]


    print("Compare xlsx vs xml")
    # should use tests.xml as the base, not the xlsx
    while row2 <= maxRow2:
        resName2 = ws2.cell(row=row2, column=2).value
        lim2 = ws2.cell(row=row2, column=3).value
        uidNum2  = ws2.cell(row=row2, column=5).value
        uom2  = ws2.cell(row=row2, column=6).value

        pat = re.compile(r'(-?\d+\.?\d*)-(-?\d+\.?\d*)')
        matchOut = pat.match(lim2)
        lslInXlsx = matchOut.group(1)
        uslInXlsx = matchOut.group(2)

        # print '{}, {}, {}'.format(resName2, lim2, uidNum2)
        if uidNum2 != None:
            if uidNum2 not in specDictXml:
                print("{} is in {}, NOT in {}".format(uidNum2, xlsxFile, xmlFile)) 
                row2 += 1
                continue
            lslInXml = specDictXml[uidNum2][3]
            uslInXml = specDictXml[uidNum2][4]
            uomInXml = specDictXml[uidNum2][5]

            if float(lslInXlsx) != float(lslInXml) or float(uslInXlsx) != float(uslInXml) or uom2 != uomInXml: 
                # print '{}, {}: lsl = {}'.format(xlsxFile, uidNum2, lslInXlsx)
                # print '{}, {}: lsl = {}'.format(xmlFile, uidNum2, lslInXml)
                # print "-"*20
                # print '{}, {}: usl = {}'.format(xlsxFile, uidNum2, uslInXlsx)
                # print '{}, {}: usl = {}'.format(xmlFile, uidNum2, uslInXml)
                # print "-"*20

                print '{:>30}, {:>15}: lsl = {:>6}, usl = {:>6}, units = {:>10}'.format(xlsxFile, uidNum2, lslInXlsx, uslInXlsx, uom2)
                print '{:>30}, {:>15}: lsl = {:>6}, usl = {:>6}, units = {:>10}'.format(xmlFile,  uidNum2, lslInXml,  uslInXml,  uomInXml)
                print "-"*20
        else:
            print "here are specs without UID yet"
            print '{}, {}, {}'.format(resName2, lim2, uidNum2)

        row2 += 1

    """

    print("Compare xml vs xlsx")

    # parse xlsx and save all specs in structure
    specDistXlsx = {}
    while row2 <= maxRow2:
        resName2 = ws2.cell(row=row2, column=2).value
        lim2 = ws2.cell(row=row2, column=3).value
        uidNum2  = ws2.cell(row=row2, column=5).value
        uom2  = ws2.cell(row=row2, column=6).value

        pat = re.compile(r'(-?\d+\.?\d*)-(-?\d+\.?\d*)')
        matchOut = pat.match(lim2)
        lslInXlsx = matchOut.group(1)
        uslInXlsx = matchOut.group(2)

        specDistXlsx[resName2] = {'lsl':lslInXlsx, \
                                  'usl':uslInXlsx, \
                                  'uidNum':uidNum2, \
                                  'unit':uom2}
        row2 += 1

    tree1 = ET.parse(xmlFile)
    root1 = tree1.getroot()
    specDictXml = getOperationName(root1)


    """

def revertAbbreviation(abrev):
    getAbrevList()

# errcount -> err
# but we can't revert err to errcount, as we may have filtErr 
# so the work around is, to generate xlsx without using abbreviations
def getAbrevList():
    abrevList = []
    with open('abbreviation list.txt', 'r') as f:
        for line in f:
            abrevList.append(' '.join(reversed(line.split())))
    return abrevList 

def getAllAncestors(elem):
	allAnces = []
	# get all ancestors
	for anc in elem.iterancestors():
		allAnces.append(anc.get('name'))
	
	return '/'.join(allAnces[::-1])

# check
def checkXmlSapflagRefer(root):
    for chm in root.getchildren():
        if chm.tag is not ET.Comment and 'spec' in chm.tag:
            # if flag_sap is 1, but reference != SP1098_PSM, report error
            curChmName = getAllAncestors(chm) + '/' + chm.get('name')
            if chm.get('flag_sap') == "1" and chm.get('reference') != 'SP1098_PSM':
                if 'Module' not in curChmName:
                    print curChmName
                    print chm.items()
                    print "\n"

            if 'Module' in curChmName and chm.get('reference') != None:
                print curChmName
                print chm.items()
                print "\n"

            if chm.get('flag_sap') == "0" and chm.get('reference') != None: 
                if 'Module' not in curChmName:
                    print curChmName
                    print chm.items()
                    print "\n"
        else:
            checkXmlSapflagRefer(chm)
    return

def checkReferences(xmlFile):
    print("+"*80)
    print("---------  checking reference errors of {}".format(xmlFile))
    tree = ET.parse(xmlFile)
    root = tree.getroot()
    checkXmlSapflagRefer(root)
    print("---------  checking reference errors done")
    print("+"*80)

if __name__ == '__main__':
    masterXml = r'C:\SystemNPI\Projects\SP1098\DteInterface\SP1098.2_P2_Limited_Launch\AIM\aim_style_3221_master_sequence.xml'
    # masterXml = 'sp1098_psm_tests.xml'
    testXlsx = '831661-07-A.xlsx'

    # testXml = r'C:\SystemNPI\Projects\SP1098\PSM2\trunk\sp1098_psm_tests_unittest.xml'
    testXml = r'C:\SystemNPI\Projects\SP1098\PSM2\trunk\sp1098_motorpack_tests.xml'
    # testXml = r'C:\SystemNPI\Projects\SP1098\PSM2\trunk\sp1098_insertionmotor_tests.xml'
    # testXml = 'sp1098_motorpack_tests.xml'
    # testXml = 'sp1098_insertionmotor_tests.xml'

    compareTwoXmls(masterXml, testXml)
    # checkReferences(testXml)

    # compareXmlXlsx(testXml, testXlsx)
